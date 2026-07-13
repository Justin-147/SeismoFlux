"""Deterministic earthquake catalog parsing and conservative physical-event deduplication."""

from __future__ import annotations

import math
import re
from bisect import bisect_left, bisect_right
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import UTC, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Final, Literal

from openpyxl import load_workbook  # type: ignore[import-untyped]
from shapely import prepare
from shapely.geometry import Point
from shapely.geometry.base import BaseGeometry

from seismoflux.data.common import (
    BEIJING_FIXED_OFFSET,
    SourceFile,
    canonical_decimal,
    canonical_json_bytes,
    normalize_text,
    read_stable_bytes,
    sha256_bytes,
    stable_token,
)

CatalogKind = Literal["m3", "m5"]
QualityScalar = object

_DATE_PATTERN: Final = re.compile(r"(?P<year>\d{4})(?P<month>\d{2})(?P<day>\d{2})")
_TIME_PATTERN: Final = re.compile(r"(?P<hour>\d{2})(?P<minute>\d{2})(?P<second>\d{2}(?:\.\d+)?)")
_CJK_PATTERN: Final = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff]")
_M5_HEADERS: Final = ("年", "月", "日", "时", "分", "秒", "纬度", "经度", "震级", "深度", "地名")
_EARTH_RADIUS_KM: Final = 6371.0088
_BASE_QUALITY_FLAGS: Final = (
    "magnitude_type_missing",
    "publication_time_assumed_origin_time",
    "source_crs_assumed_wgs84",
    "source_timezone_assumed_fixed_utc_plus_08",
)


@dataclass(frozen=True, slots=True)
class DedupThresholds:
    """Auditable joint thresholds for a possible cross-catalog duplicate."""

    max_time_delta_seconds: float
    max_distance_km: float
    max_magnitude_delta: float

    def __post_init__(self) -> None:
        for name, value in (
            ("max_time_delta_seconds", self.max_time_delta_seconds),
            ("max_distance_km", self.max_distance_km),
            ("max_magnitude_delta", self.max_magnitude_delta),
        ):
            if not math.isfinite(value) or value < 0:
                raise ValueError(f"{name} must be a finite non-negative number")


DEFAULT_AUTO_THRESHOLDS: Final = DedupThresholds(5.0, 5.0, 0.1)
DEFAULT_REVIEW_THRESHOLDS: Final = DedupThresholds(300.0, 50.0, 0.5)


@dataclass(frozen=True, slots=True)
class EarthquakeEvent:
    """One physical event after accepted same- and cross-catalog deduplication."""

    event_id: str
    origin_time_utc: datetime
    available_at: datetime
    origin_time_local: datetime
    longitude: float
    latitude: float
    depth_km: float | None
    magnitude: float
    magnitude_type: str | None
    place: str | None
    catalog_sources: tuple[str, ...]
    inside_study_area: bool
    dedup_confidence: str
    anchor_source_record_id: str
    quality_flags: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class EarthquakeSourceRecord:
    """A normalized row that remains traceable to immutable source bytes."""

    source_record_id: str
    duplicate_group_id: str
    source_id: str
    source_file: str
    source_file_sha256: str
    source_row: int
    raw_record_sha256: str
    origin_time_raw: str
    origin_time_local: datetime
    origin_time_utc: datetime
    available_at: datetime
    longitude_raw: str
    latitude_raw: str
    depth_raw: str
    magnitude_raw: str
    longitude: float
    latitude: float
    depth_km: float | None
    magnitude: float
    magnitude_type: str | None
    place_raw: str | None
    place: str | None
    fixed_field_raw: str | None
    inside_study_area: bool
    normalization_flags: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class DedupCandidate:
    """A reviewable edge between one M3 group and one M5 group."""

    candidate_id: str
    m3_group_id: str
    m5_group_id: str
    m3_source_record_ids: tuple[str, ...]
    m5_source_record_ids: tuple[str, ...]
    time_delta_seconds: float
    distance_km: float
    magnitude_delta: float
    exact_match: bool
    auto_eligible: bool
    m3_auto_degree: int
    m5_auto_degree: int
    decision: str
    merged_event_id: str | None


@dataclass(frozen=True, slots=True)
class EarthquakeParseResult:
    """All deterministic normalized rows, events, candidates, and scalar QA evidence."""

    events: tuple[EarthquakeEvent, ...]
    source_records: tuple[EarthquakeSourceRecord, ...]
    dedup_candidates: tuple[DedupCandidate, ...]
    quality: dict[str, QualityScalar]


@dataclass(frozen=True, slots=True)
class _ParsedRow:
    kind: CatalogKind
    source: SourceFile
    source_row: int
    raw_record_sha256: str
    semantic_digest: str
    origin_time_raw: str
    origin_time_local: datetime
    origin_time_utc: datetime
    longitude_raw: str
    latitude_raw: str
    depth_raw: str
    magnitude_raw: str
    longitude_decimal: Decimal
    latitude_decimal: Decimal
    depth_decimal: Decimal | None
    magnitude_decimal: Decimal
    place_raw: str | None
    place: str | None
    fixed_field_raw: str | None
    normalization_flags: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class _SourceGroup:
    kind: CatalogKind
    group_id: str
    parsed: _ParsedRow
    records: tuple[EarthquakeSourceRecord, ...]


@dataclass(frozen=True, slots=True)
class _CandidateDraft:
    m3: _SourceGroup
    m5: _SourceGroup
    time_delta_seconds: float
    distance_km: float
    magnitude_delta: float
    exact_match: bool
    auto_eligible: bool


def _decimal(value: object, *, field: str, locator: str) -> Decimal:
    try:
        return Decimal(canonical_decimal(value))
    except ValueError as exc:
        raise ValueError(f"{locator}: invalid {field}: {value!r}") from exc


def _integer(value: object, *, field: str, locator: str) -> int:
    decimal_value = _decimal(value, field=field, locator=locator)
    if decimal_value != decimal_value.to_integral_value():
        raise ValueError(f"{locator}: {field} must be an integer: {value!r}")
    return int(decimal_value)


def _fixed_offset_datetime(
    year: int,
    month: int,
    day: int,
    hour: int,
    minute: int,
    second: Decimal,
    *,
    allow_second_60: bool,
    locator: str,
) -> datetime:
    invalid_second = second < 0 or second > 60
    invalid_second = invalid_second or (not allow_second_60 and second == 60)
    if invalid_second:
        qualifier = "0..60" if allow_second_60 else "0..<60"
        raise ValueError(f"{locator}: second must be in {qualifier}: {second}")

    microseconds = second * Decimal(1_000_000)
    if microseconds != microseconds.to_integral_value():
        raise ValueError(f"{locator}: sub-microsecond origin time is unsupported: {second}")
    try:
        base = datetime(year, month, day, hour, minute, tzinfo=BEIJING_FIXED_OFFSET)
    except ValueError as exc:
        raise ValueError(f"{locator}: invalid origin date or time") from exc
    return base + timedelta(microseconds=int(microseconds))


def _validate_physical_values(
    latitude: Decimal,
    longitude: Decimal,
    magnitude: Decimal,
    *,
    locator: str,
) -> None:
    if not Decimal(-90) <= latitude <= Decimal(90):
        raise ValueError(f"{locator}: latitude outside [-90, 90]: {latitude}")
    if not Decimal(-180) <= longitude <= Decimal(180):
        raise ValueError(f"{locator}: longitude outside [-180, 180]: {longitude}")
    if magnitude < 0:
        raise ValueError(f"{locator}: magnitude must be non-negative: {magnitude}")


def _repair_place(raw_value: str | None) -> tuple[str | None, bool]:
    if raw_value is None or not raw_value.strip():
        return None, False
    stripped = raw_value.strip()
    if _CJK_PATTERN.search(stripped):
        return normalize_text(stripped), False
    if all(ord(character) <= 255 for character in stripped):
        try:
            repaired = stripped.encode("latin-1", errors="strict").decode(
                "gb18030", errors="strict"
            )
        except UnicodeError:
            repaired = ""
        if repaired and _CJK_PATTERN.search(repaired):
            return normalize_text(repaired), True
    return normalize_text(stripped), False


def _depth_value(raw_depth: Decimal, flags: list[str]) -> Decimal | None:
    if raw_depth < 0:
        flags.append("negative_depth_sentinel")
        return None
    return raw_depth


def _parse_m3_catalog(source: SourceFile, payload: bytes) -> list[_ParsedRow]:
    try:
        text = payload.decode("utf-8", errors="strict")
    except UnicodeDecodeError as exc:
        raise ValueError(f"{source.relative_path}: expected strict UTF-8 bytes") from exc

    parsed: list[_ParsedRow] = []
    for line_number, line in enumerate(text.splitlines(), start=1):
        locator = f"{source.relative_path}:line {line_number}"
        parts = line.split(maxsplit=7)
        if len(parts) < 7:
            raise ValueError(f"{locator}: expected at least seven fields")
        date_raw, time_raw, latitude_raw, longitude_raw, magnitude_raw, depth_raw, fixed = parts[:7]
        place_raw = parts[7] if len(parts) == 8 else None

        date_match = _DATE_PATTERN.fullmatch(date_raw)
        time_match = _TIME_PATTERN.fullmatch(time_raw)
        if date_match is None or time_match is None:
            raise ValueError(f"{locator}: invalid compact origin time")
        second = _decimal(time_match.group("second"), field="second", locator=locator)
        local_time = _fixed_offset_datetime(
            int(date_match.group("year")),
            int(date_match.group("month")),
            int(date_match.group("day")),
            int(time_match.group("hour")),
            int(time_match.group("minute")),
            second,
            allow_second_60=False,
            locator=locator,
        )

        latitude = _decimal(latitude_raw, field="latitude", locator=locator)
        longitude = _decimal(longitude_raw, field="longitude", locator=locator)
        magnitude = _decimal(magnitude_raw, field="magnitude", locator=locator)
        raw_depth = _decimal(depth_raw, field="depth", locator=locator)
        _validate_physical_values(latitude, longitude, magnitude, locator=locator)

        flags = list(_BASE_QUALITY_FLAGS)
        depth = _depth_value(raw_depth, flags)
        place, repaired = _repair_place(place_raw)
        if repaired:
            flags.append("place_repaired_latin1_gb18030")
        if place is None:
            flags.append("place_missing")
        if magnitude < 3:
            flags.append("magnitude_below_nominal_m3")
        if fixed != "000":
            flags.append("fixed_field_variant")

        semantic_fields = {
            "date": date_raw,
            "depth": canonical_decimal(raw_depth),
            "fixed": fixed,
            "latitude": canonical_decimal(latitude),
            "longitude": canonical_decimal(longitude),
            "magnitude": canonical_decimal(magnitude),
            "place": place_raw,
            "time": time_raw,
        }
        semantic_bytes = canonical_json_bytes(semantic_fields)
        parsed.append(
            _ParsedRow(
                kind="m3",
                source=source,
                source_row=line_number,
                raw_record_sha256=sha256_bytes(line.encode("utf-8")),
                semantic_digest=sha256_bytes(semantic_bytes),
                origin_time_raw=f"{date_raw} {time_raw}",
                origin_time_local=local_time,
                origin_time_utc=local_time.astimezone(UTC),
                longitude_raw=longitude_raw,
                latitude_raw=latitude_raw,
                depth_raw=depth_raw,
                magnitude_raw=magnitude_raw,
                longitude_decimal=longitude,
                latitude_decimal=latitude,
                depth_decimal=depth,
                magnitude_decimal=magnitude,
                place_raw=place_raw,
                place=place,
                fixed_field_raw=fixed,
                normalization_flags=tuple(sorted(set(flags))),
            )
        )
    return parsed


def _parse_m5_catalog(source: SourceFile, payload: bytes) -> list[_ParsedRow]:
    try:
        workbook = load_workbook(BytesIO(payload), read_only=True, data_only=False)
    except Exception as exc:
        raise ValueError(f"{source.relative_path}: invalid XLSX workbook") from exc

    try:
        if len(workbook.sheetnames) != 1:
            raise ValueError(f"{source.relative_path}: expected exactly one worksheet")
        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(min_col=1, max_col=len(_M5_HEADERS), values_only=True)
        try:
            header_values = next(rows)
        except StopIteration as exc:
            raise ValueError(f"{source.relative_path}: workbook is empty") from exc
        headers = tuple(normalize_text(value) for value in header_values)
        if headers != _M5_HEADERS:
            raise ValueError(f"{source.relative_path}: unexpected headers: {headers!r}")

        parsed: list[_ParsedRow] = []
        for row_number, values in enumerate(rows, start=2):
            locator = f"{source.relative_path}:row {row_number}"
            if any(value is None or normalize_text(value) is None for value in values):
                raise ValueError(f"{locator}: all eleven cells are required")
            year, month, day, hour, minute = (
                _integer(values[index], field=_M5_HEADERS[index], locator=locator)
                for index in range(5)
            )
            second = _decimal(values[5], field="秒", locator=locator)
            local_time = _fixed_offset_datetime(
                year,
                month,
                day,
                hour,
                minute,
                second,
                allow_second_60=True,
                locator=locator,
            )

            latitude = _decimal(values[6], field="纬度", locator=locator)
            longitude = _decimal(values[7], field="经度", locator=locator)
            magnitude = _decimal(values[8], field="震级", locator=locator)
            raw_depth = _decimal(values[9], field="深度", locator=locator)
            _validate_physical_values(latitude, longitude, magnitude, locator=locator)

            flags = list(_BASE_QUALITY_FLAGS)
            depth = _depth_value(raw_depth, flags)
            place_raw = str(values[10])
            place = normalize_text(place_raw)
            if place is None:
                flags.append("place_missing")
            if second == 60:
                flags.append("second_60_carried")
            if hour == 0 and minute == 0 and second == 0:
                flags.append("possible_day_level_time")
            if raw_depth == 10:
                flags.append("possible_default_depth_10")
            if magnitude < 5:
                flags.append("magnitude_below_nominal_m5")

            numeric_raw = tuple(canonical_decimal(value) for value in values[:10])
            semantic_fields = {
                "day": numeric_raw[2],
                "depth": numeric_raw[9],
                "hour": numeric_raw[3],
                "latitude": numeric_raw[6],
                "longitude": numeric_raw[7],
                "magnitude": numeric_raw[8],
                "minute": numeric_raw[4],
                "month": numeric_raw[1],
                "place": place_raw,
                "second": numeric_raw[5],
                "year": numeric_raw[0],
            }
            semantic_bytes = canonical_json_bytes(semantic_fields)
            origin_time_raw = (
                f"{year:04d}-{month:02d}-{day:02d} "
                f"{hour:02d}:{minute:02d}:{canonical_decimal(second)}"
            )
            parsed.append(
                _ParsedRow(
                    kind="m5",
                    source=source,
                    source_row=row_number,
                    raw_record_sha256=sha256_bytes(semantic_bytes),
                    semantic_digest=sha256_bytes(semantic_bytes),
                    origin_time_raw=origin_time_raw,
                    origin_time_local=local_time,
                    origin_time_utc=local_time.astimezone(UTC),
                    longitude_raw=numeric_raw[7],
                    latitude_raw=numeric_raw[6],
                    depth_raw=numeric_raw[9],
                    magnitude_raw=numeric_raw[8],
                    longitude_decimal=longitude,
                    latitude_decimal=latitude,
                    depth_decimal=depth,
                    magnitude_decimal=magnitude,
                    place_raw=place_raw,
                    place=place,
                    fixed_field_raw=None,
                    normalization_flags=tuple(sorted(set(flags))),
                )
            )
        return parsed
    finally:
        workbook.close()


def _group_records(
    parsed_rows: list[_ParsedRow], study_area: BaseGeometry
) -> tuple[list[EarthquakeSourceRecord], list[_SourceGroup]]:
    grouped: defaultdict[str, list[_ParsedRow]] = defaultdict(list)
    for row in parsed_rows:
        grouped[row.semantic_digest].append(row)

    source_records: list[EarthquakeSourceRecord] = []
    source_groups: list[_SourceGroup] = []
    for semantic_digest in sorted(grouped):
        rows = sorted(grouped[semantic_digest], key=lambda item: item.source_row)
        first = rows[0]
        group_id = stable_token("eqgrp", first.kind, first.source.source_id, semantic_digest)
        records: list[EarthquakeSourceRecord] = []
        duplicate = len(rows) > 1
        for duplicate_ordinal, row in enumerate(rows, start=1):
            flags = list(row.normalization_flags)
            if duplicate:
                flags.append("same_source_exact_duplicate_collapsed")
            source_record_id = stable_token(
                "eqsrc", row.source.source_id, semantic_digest, duplicate_ordinal
            )
            inside = bool(
                study_area.covers(Point(float(row.longitude_decimal), float(row.latitude_decimal)))
            )
            record = EarthquakeSourceRecord(
                source_record_id=source_record_id,
                duplicate_group_id=group_id,
                source_id=row.source.source_id,
                source_file=row.source.relative_path,
                source_file_sha256=row.source.sha256,
                source_row=row.source_row,
                raw_record_sha256=row.raw_record_sha256,
                origin_time_raw=row.origin_time_raw,
                origin_time_local=row.origin_time_local,
                origin_time_utc=row.origin_time_utc,
                available_at=row.origin_time_utc,
                longitude_raw=row.longitude_raw,
                latitude_raw=row.latitude_raw,
                depth_raw=row.depth_raw,
                magnitude_raw=row.magnitude_raw,
                longitude=float(row.longitude_decimal),
                latitude=float(row.latitude_decimal),
                depth_km=None if row.depth_decimal is None else float(row.depth_decimal),
                magnitude=float(row.magnitude_decimal),
                magnitude_type=None,
                place_raw=row.place_raw,
                place=row.place,
                fixed_field_raw=row.fixed_field_raw,
                inside_study_area=inside,
                normalization_flags=tuple(sorted(set(flags))),
            )
            records.append(record)
            source_records.append(record)
        source_groups.append(
            _SourceGroup(
                kind=first.kind,
                group_id=group_id,
                parsed=first,
                records=tuple(records),
            )
        )
    return source_records, source_groups


def _great_circle_km(left: _ParsedRow, right: _ParsedRow) -> float:
    if (
        left.latitude_decimal == right.latitude_decimal
        and left.longitude_decimal == right.longitude_decimal
    ):
        return 0.0
    latitude_1 = math.radians(float(left.latitude_decimal))
    longitude_1 = math.radians(float(left.longitude_decimal))
    latitude_2 = math.radians(float(right.latitude_decimal))
    longitude_2 = math.radians(float(right.longitude_decimal))
    latitude_delta = latitude_2 - latitude_1
    longitude_delta = longitude_2 - longitude_1
    haversine = math.sin(latitude_delta / 2) ** 2 + (
        math.cos(latitude_1) * math.cos(latitude_2) * math.sin(longitude_delta / 2) ** 2
    )
    distance_km = 2 * _EARTH_RADIUS_KM * math.asin(min(1.0, math.sqrt(haversine)))
    return round(distance_km, 9)


def _candidate_drafts(
    m3_groups: list[_SourceGroup],
    m5_groups: list[_SourceGroup],
    auto: DedupThresholds,
    review: DedupThresholds,
) -> list[_CandidateDraft]:
    sorted_m5 = sorted(
        m5_groups,
        key=lambda group: (group.parsed.origin_time_utc, group.group_id),
    )
    m5_times = [group.parsed.origin_time_utc for group in sorted_m5]
    review_magnitude = Decimal(str(review.max_magnitude_delta))
    auto_magnitude = Decimal(str(auto.max_magnitude_delta))
    review_window = timedelta(seconds=review.max_time_delta_seconds)

    candidates: list[_CandidateDraft] = []
    for m3_group in sorted(
        m3_groups,
        key=lambda group: (group.parsed.origin_time_utc, group.group_id),
    ):
        origin = m3_group.parsed.origin_time_utc
        start = bisect_left(m5_times, origin - review_window)
        stop = bisect_right(m5_times, origin + review_window)
        for m5_group in sorted_m5[start:stop]:
            time_delta = abs((origin - m5_group.parsed.origin_time_utc).total_seconds())
            magnitude_delta_decimal = abs(
                m3_group.parsed.magnitude_decimal - m5_group.parsed.magnitude_decimal
            )
            if magnitude_delta_decimal > review_magnitude:
                continue
            distance = _great_circle_km(m3_group.parsed, m5_group.parsed)
            if distance > review.max_distance_km:
                continue
            magnitude_delta = float(magnitude_delta_decimal)
            auto_eligible = (
                time_delta <= auto.max_time_delta_seconds
                and distance <= auto.max_distance_km
                and magnitude_delta_decimal <= auto_magnitude
            )
            exact_match = (
                time_delta == 0
                and m3_group.parsed.latitude_decimal == m5_group.parsed.latitude_decimal
                and m3_group.parsed.longitude_decimal == m5_group.parsed.longitude_decimal
                and m3_group.parsed.magnitude_decimal == m5_group.parsed.magnitude_decimal
            )
            candidates.append(
                _CandidateDraft(
                    m3=m3_group,
                    m5=m5_group,
                    time_delta_seconds=time_delta,
                    distance_km=distance,
                    magnitude_delta=magnitude_delta,
                    exact_match=exact_match,
                    auto_eligible=auto_eligible,
                )
            )
    return sorted(
        candidates,
        key=lambda candidate: (
            candidate.time_delta_seconds,
            candidate.distance_km,
            candidate.magnitude_delta,
            candidate.m3.group_id,
            candidate.m5.group_id,
        ),
    )


def _record_sort_key(
    record: EarthquakeSourceRecord, m3_source_id: str, m5_source_id: str
) -> tuple[int, int, str]:
    priority = 0 if record.source_id == m5_source_id else 1
    if record.source_id not in {m3_source_id, m5_source_id}:
        priority = 2
    return priority, record.source_row, record.source_record_id


def _build_event(
    groups: tuple[_SourceGroup, ...],
    confidence: str,
    m3_source_id: str,
    m5_source_id: str,
) -> EarthquakeEvent:
    ordered_groups = sorted(
        groups,
        key=lambda group: (0 if group.kind == "m5" else 1, group.group_id),
    )
    anchor_group = ordered_groups[0]
    anchor = anchor_group.records[0]
    records = sorted(
        (record for group in ordered_groups for record in group.records),
        key=lambda record: _record_sort_key(record, m3_source_id, m5_source_id),
    )
    flags = {flag for record in records for flag in record.normalization_flags}
    if len(groups) == 2:
        flags.add("cross_catalog_merged")
    if confidence in {"ambiguous_unmerged", "review_unmerged"}:
        flags.add(confidence)
    return EarthquakeEvent(
        event_id=stable_token("eqv1", anchor.source_id, anchor.source_record_id),
        origin_time_utc=anchor.origin_time_utc,
        available_at=anchor.origin_time_utc,
        origin_time_local=anchor.origin_time_local,
        longitude=anchor.longitude,
        latitude=anchor.latitude,
        depth_km=anchor.depth_km,
        magnitude=anchor.magnitude,
        magnitude_type=None,
        place=anchor.place,
        catalog_sources=tuple(record.source_record_id for record in records),
        inside_study_area=anchor.inside_study_area,
        dedup_confidence=confidence,
        anchor_source_record_id=anchor.source_record_id,
        quality_flags=tuple(sorted(flags)),
    )


def _validate_inputs(
    m3_file: SourceFile,
    m5_file: SourceFile,
    study_area: BaseGeometry,
    auto: DedupThresholds,
    review: DedupThresholds,
) -> None:
    if m3_file.source_id == m5_file.source_id:
        raise ValueError("M3 and M5 catalogs must use different source IDs")
    if study_area.is_empty or not study_area.is_valid:
        raise ValueError("study_area must be a non-empty valid polygon")
    if study_area.geom_type not in {"Polygon", "MultiPolygon"}:
        raise ValueError("study_area must be a Polygon or MultiPolygon")
    if (
        auto.max_time_delta_seconds > review.max_time_delta_seconds
        or auto.max_distance_km > review.max_distance_km
        or auto.max_magnitude_delta > review.max_magnitude_delta
    ):
        raise ValueError("automatic dedup thresholds must be nested within review thresholds")


def parse_earthquake_catalogs(
    m3_file: SourceFile,
    m5_file: SourceFile,
    study_area: BaseGeometry,
    auto: DedupThresholds = DEFAULT_AUTO_THRESHOLDS,
    review: DedupThresholds = DEFAULT_REVIEW_THRESHOLDS,
) -> EarthquakeParseResult:
    """Parse, normalize, and conservatively deduplicate the two locked catalogs.

    Source rows are never mutated. Exact same-source rows are grouped first. Cross-source
    rows merge only when they satisfy the automatic joint rule and both endpoint degrees in
    the automatic candidate graph are one. Wider review candidates are retained but never
    merged automatically.
    """

    _validate_inputs(m3_file, m5_file, study_area, auto, review)
    prepare(study_area)
    m3_parsed = _parse_m3_catalog(m3_file, read_stable_bytes(m3_file))
    m5_parsed = _parse_m5_catalog(m5_file, read_stable_bytes(m5_file))
    m3_records, m3_groups = _group_records(m3_parsed, study_area)
    m5_records, m5_groups = _group_records(m5_parsed, study_area)

    drafts = _candidate_drafts(m3_groups, m5_groups, auto, review)
    m3_auto_degree = Counter(draft.m3.group_id for draft in drafts if draft.auto_eligible)
    m5_auto_degree = Counter(draft.m5.group_id for draft in drafts if draft.auto_eligible)
    accepted = [
        draft
        for draft in drafts
        if draft.auto_eligible
        and m3_auto_degree[draft.m3.group_id] == 1
        and m5_auto_degree[draft.m5.group_id] == 1
    ]

    matched_group_ids = {
        group_id for draft in accepted for group_id in (draft.m3.group_id, draft.m5.group_id)
    }
    ambiguous_group_ids = {
        group_id
        for draft in drafts
        if draft.auto_eligible
        and (m3_auto_degree[draft.m3.group_id] != 1 or m5_auto_degree[draft.m5.group_id] != 1)
        for group_id in (draft.m3.group_id, draft.m5.group_id)
    }
    review_group_ids = {
        group_id
        for draft in drafts
        if not draft.auto_eligible
        for group_id in (draft.m3.group_id, draft.m5.group_id)
    }

    events: list[EarthquakeEvent] = []
    group_to_event_id: dict[str, str] = {}
    for draft in accepted:
        confidence = "exact" if draft.exact_match else "tight_unique"
        event = _build_event((draft.m3, draft.m5), confidence, m3_file.source_id, m5_file.source_id)
        events.append(event)
        group_to_event_id[draft.m3.group_id] = event.event_id
        group_to_event_id[draft.m5.group_id] = event.event_id

    for group in sorted(
        [*m3_groups, *m5_groups],
        key=lambda item: (item.parsed.origin_time_utc, item.kind, item.group_id),
    ):
        if group.group_id in matched_group_ids:
            continue
        if group.group_id in ambiguous_group_ids:
            confidence = "ambiguous_unmerged"
        elif group.group_id in review_group_ids:
            confidence = "review_unmerged"
        elif len(group.records) > 1:
            confidence = "exact"
        else:
            confidence = "unmatched"
        event = _build_event((group,), confidence, m3_file.source_id, m5_file.source_id)
        events.append(event)
        group_to_event_id[group.group_id] = event.event_id

    candidates: list[DedupCandidate] = []
    for draft in drafts:
        m3_degree = m3_auto_degree[draft.m3.group_id]
        m5_degree = m5_auto_degree[draft.m5.group_id]
        auto_merged = draft.auto_eligible and m3_degree == 1 and m5_degree == 1
        if auto_merged:
            decision = "auto_merged"
            merged_event_id = group_to_event_id[draft.m3.group_id]
        elif draft.auto_eligible:
            decision = "ambiguous_unmerged"
            merged_event_id = None
        else:
            decision = "review_unmerged"
            merged_event_id = None
        candidates.append(
            DedupCandidate(
                candidate_id=stable_token("eqcand", draft.m3.group_id, draft.m5.group_id),
                m3_group_id=draft.m3.group_id,
                m5_group_id=draft.m5.group_id,
                m3_source_record_ids=tuple(record.source_record_id for record in draft.m3.records),
                m5_source_record_ids=tuple(record.source_record_id for record in draft.m5.records),
                time_delta_seconds=draft.time_delta_seconds,
                distance_km=draft.distance_km,
                magnitude_delta=draft.magnitude_delta,
                exact_match=draft.exact_match,
                auto_eligible=draft.auto_eligible,
                m3_auto_degree=m3_degree,
                m5_auto_degree=m5_degree,
                decision=decision,
                merged_event_id=merged_event_id,
            )
        )

    source_records = sorted(
        [*m3_records, *m5_records],
        key=lambda record: (
            record.origin_time_utc,
            record.source_id,
            record.source_row,
            record.source_record_id,
        ),
    )
    events.sort(key=lambda event: (event.origin_time_utc, event.event_id))
    candidates.sort(key=lambda candidate: candidate.candidate_id)

    same_source_duplicate_groups = sum(len(group.records) > 1 for group in [*m3_groups, *m5_groups])
    same_source_duplicate_records_collapsed = sum(
        len(group.records) - 1 for group in [*m3_groups, *m5_groups]
    )
    depths = [record.depth_km for record in source_records if record.depth_km is not None]
    quality: dict[str, QualityScalar] = {
        "schema_version": "earthquake-v1",
        "timezone": "UTC+08:00",
        "m3_input_records": len(m3_parsed),
        "m5_input_records": len(m5_parsed),
        "source_records": len(source_records),
        "origin_time_min_utc": min(record.origin_time_utc for record in source_records).isoformat(),
        "origin_time_max_utc": max(record.origin_time_utc for record in source_records).isoformat(),
        "available_at_min_utc": min(record.available_at for record in source_records).isoformat(),
        "available_at_max_utc": max(record.available_at for record in source_records).isoformat(),
        "longitude_min": min(record.longitude for record in source_records),
        "longitude_max": max(record.longitude for record in source_records),
        "latitude_min": min(record.latitude for record in source_records),
        "latitude_max": max(record.latitude for record in source_records),
        "depth_km_min": min(depths) if depths else None,
        "depth_km_max": max(depths) if depths else None,
        "magnitude_min": min(record.magnitude for record in source_records),
        "magnitude_max": max(record.magnitude for record in source_records),
        "depth_missing_records": sum(record.depth_km is None for record in source_records),
        "magnitude_type_missing_records": sum(
            record.magnitude_type is None for record in source_records
        ),
        "publication_time_assumption_records": sum(
            "publication_time_assumed_origin_time" in record.normalization_flags
            for record in source_records
        ),
        "same_source_duplicate_groups": same_source_duplicate_groups,
        "same_source_duplicate_records_collapsed": (same_source_duplicate_records_collapsed),
        "cross_review_candidates": len(candidates),
        "cross_exact_candidates": sum(candidate.exact_match for candidate in candidates),
        "cross_auto_eligible_candidates": sum(candidate.auto_eligible for candidate in candidates),
        "cross_auto_merged_candidates": len(accepted),
        "cross_auto_ambiguous_edges": sum(
            candidate.decision == "ambiguous_unmerged" for candidate in candidates
        ),
        "events": len(events),
        "inside_study_area_events": sum(event.inside_study_area for event in events),
        "negative_depth_records": sum(
            "negative_depth_sentinel" in record.normalization_flags for record in source_records
        ),
        "repaired_place_records": sum(
            "place_repaired_latin1_gb18030" in record.normalization_flags
            for record in source_records
        ),
        "m3_below_3_records": sum(
            "magnitude_below_nominal_m3" in record.normalization_flags for record in source_records
        ),
        "m5_second_60_records": sum(
            "second_60_carried" in record.normalization_flags for record in source_records
        ),
        "m5_possible_day_level_time_records": sum(
            "possible_day_level_time" in record.normalization_flags for record in source_records
        ),
        "m5_possible_default_depth_10_records": sum(
            "possible_default_depth_10" in record.normalization_flags for record in source_records
        ),
        "m5_below_5_records": sum(
            "magnitude_below_nominal_m5" in record.normalization_flags for record in source_records
        ),
        "place_missing_records": sum(
            "place_missing" in record.normalization_flags for record in source_records
        ),
    }
    return EarthquakeParseResult(
        events=tuple(events),
        source_records=tuple(source_records),
        dedup_candidates=tuple(candidates),
        quality=quality,
    )

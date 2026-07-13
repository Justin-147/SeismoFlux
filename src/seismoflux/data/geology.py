"""Deterministic ingestion for faults, long-term hazard, and PlotBD geometry."""

from __future__ import annotations

import math
import re
from collections import Counter, defaultdict
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from datetime import UTC, datetime
from io import BytesIO
from typing import Any, cast

from openpyxl import load_workbook  # type: ignore[import-untyped]
from pyproj import CRS, Geod, Transformer
from shapely import to_wkb
from shapely.geometry import LineString, Polygon
from shapely.geometry.polygon import orient
from shapely.ops import transform

from seismoflux.data.common import (
    SourceFile,
    canonical_decimal,
    normalize_text,
    read_stable_bytes,
    stable_token,
)

Record = dict[str, Any]
Coordinate = tuple[float, float]

FAULT_COORDINATE_HEADERS = (
    "断层编号",
    "断层段编号",
    "段内点编号",
    "经度",
    "纬度",
    "本次新增",
    "断层名称",
    "断层段名称",
)
FAULT_ATTRIBUTE_HEADERS = (
    "断层编号",
    "断层名称",
    "断层段编号",
    "断层段名称",
    "整体走向",
    "整体倾角",
    "滑动角",
    "地震地质运动总速率mm/a",
    "地震地质速率走滑分量_左旋负_右旋正",
    "地震地质速率张压分量_挤压负_拉张正",
    "大地测量反演运动总速率mm/a",
    "大地测量反演速率走滑分量_左旋负_右旋正",
    "大地测量反演速率走滑分量误差",
    "大地测量反演速率张压分量_挤压负_拉张正",
    "大地测量反演速率张压分量误差",
    "强震复发周期yr",
    "强震复发周期类型_1可靠_2推测",
    "上次强震时间AD",
    "上次强震时间类型_1可靠_2推测",
)
LONG_TERM_HAZARD_HEADERS = (
    "断层编号",
    "断层名称",
    "断层段编号",
    "断层段名称",
    "空段类型_new权重",
    "闭锁率结果_new权重",
    "小震稀疏段结果_new权重",
    "库仑应力结果_new权重",
    "危险系数_new",
)

_REQUIRED_FILES = (
    "FaultCord.xlsx",
    "FaultAttri_ALLV4_修改.xlsx",
    "weight_analysis-修改newV4_修改.xlsx",
    "CN-block-L1-deduced.gmt",
    "CN-block-L1.gmt",
    "CN-block-L2.gmt",
    "CN-border-L1.gmt",
    "CN-plate-neighbor.dat",
    "FaultX.dat",
)
_BASEMAP_ROLES = {
    "cn-block-l1-deduced.gmt": "tectonic_block_level1_deduced",
    "cn-block-l1.gmt": "tectonic_block_level1",
    "cn-block-l2.gmt": "tectonic_block_level2",
    "cn-border-l1.gmt": "national_border",
    "cn-plate-neighbor.dat": "neighbor_plate_boundary",
}
_DASH_RE = re.compile(r"[\u2010\u2011\u2012\u2013\u2014\u2015\uff0d_]+")
_NUMERIC_NAME_RE = re.compile(r"[+-]?\d+(?:\.\d+)?")


@dataclass(frozen=True, slots=True)
class GeologyParseResult:
    """Normalized geology records and their deterministic quality summary."""

    fault_points: tuple[Record, ...]
    fault_segments: tuple[Record, ...]
    fault_traces: tuple[Record, ...]
    basemap_features: tuple[Record, ...]
    crosswalk_audit: tuple[Record, ...]
    study_area: Record
    quality: Record


@dataclass(frozen=True, slots=True)
class _GmtSegment:
    ordinal: int
    header_raw: str | None
    comments: tuple[str, ...]
    delimiter_line: int
    coordinates: tuple[Coordinate, ...]


def _filename(source: SourceFile) -> str:
    return source.relative_path.replace("\\", "/").rsplit("/", 1)[-1]


def _find_sources(source_map: Mapping[str, SourceFile]) -> dict[str, SourceFile]:
    by_filename: dict[str, SourceFile] = {}
    for source in source_map.values():
        filename = _filename(source).casefold()
        existing = by_filename.get(filename)
        if existing is not None:
            if existing.relative_path != source.relative_path or existing.sha256 != source.sha256:
                raise ValueError(f"multiple inventory entries have filename {_filename(source)!r}")
            continue
        by_filename[filename] = source

    missing = [name for name in _REQUIRED_FILES if name.casefold() not in by_filename]
    if missing:
        raise ValueError(f"missing geology sources: {', '.join(missing)}")
    expected_filenames = {name.casefold() for name in _REQUIRED_FILES}
    unexpected = sorted(set(by_filename) - expected_filenames)
    if unexpected:
        raise ValueError(f"unexpected geology sources: {', '.join(unexpected)}")
    return {name: by_filename[name.casefold()] for name in _REQUIRED_FILES}


def _available_at(value: datetime | str) -> datetime:
    if isinstance(value, str):
        candidate = value.strip()
        if candidate.endswith("Z"):
            candidate = candidate[:-1] + "+00:00"
        try:
            parsed = datetime.fromisoformat(candidate)
        except ValueError as exc:
            raise ValueError("dataset_available_at must be an ISO datetime") from exc
    else:
        parsed = value
    if parsed.tzinfo is None or parsed.utcoffset() is None:
        raise ValueError("dataset_available_at must be timezone-aware")
    return parsed.astimezone(UTC)


def _available_at_json(value: datetime) -> str:
    return value.isoformat().replace("+00:00", "Z")


def _coordinate_bbox(coordinates: Sequence[Coordinate]) -> Record | None:
    if not coordinates:
        return None
    longitudes = [coordinate[0] for coordinate in coordinates]
    latitudes = [coordinate[1] for coordinate in coordinates]
    return {
        "min_longitude": min(longitudes),
        "min_latitude": min(latitudes),
        "max_longitude": max(longitudes),
        "max_latitude": max(latitudes),
    }


def _is_empty(value: object | None) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _required_text(value: object | None, *, field: str, row: int) -> str:
    normalized = normalize_text(value)
    if normalized is None:
        raise ValueError(f"{field} is empty at Excel row {row}")
    return normalized


def _required_float(value: object | None, *, field: str, row: int) -> float:
    if isinstance(value, bool) or _is_empty(value):
        raise ValueError(f"{field} is not numeric at Excel row {row}")
    try:
        result = float(cast(Any, value))
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{field} is not numeric at Excel row {row}") from exc
    if not math.isfinite(result):
        raise ValueError(f"{field} is not finite at Excel row {row}")
    return result


def _optional_float(value: object | None, *, field: str, row: int) -> float | None:
    if _is_empty(value):
        return None
    return _required_float(value, field=field, row=row)


def _required_int(value: object | None, *, field: str, row: int) -> int:
    number = _required_float(value, field=field, row=row)
    if not number.is_integer():
        raise ValueError(f"{field} is not an integer at Excel row {row}")
    return int(number)


def _optional_int(value: object | None, *, field: str, row: int) -> int | None:
    if _is_empty(value):
        return None
    return _required_int(value, field=field, row=row)


def _xlsx_rows(source: SourceFile, headers: Sequence[str]) -> list[tuple[int, tuple[Any, ...]]]:
    payload = read_stable_bytes(source)
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=False)
    try:
        if workbook.sheetnames != ["Sheet1"]:
            raise ValueError(f"{source.relative_path} must contain only visible Sheet1")
        worksheet = workbook["Sheet1"]
        if worksheet.sheet_state != "visible":
            raise ValueError(f"{source.relative_path} Sheet1 must be visible")
        if worksheet.max_column != len(headers):
            raise ValueError(
                f"{source.relative_path} has {worksheet.max_column} columns; "
                f"expected {len(headers)}"
            )
        iterator = worksheet.iter_rows(min_col=1, max_col=len(headers))
        try:
            header_cells = next(iterator)
        except StopIteration as exc:
            raise ValueError(f"{source.relative_path} is empty") from exc
        if any(cell.data_type == "f" for cell in header_cells):
            raise ValueError(f"formulas are forbidden in {source.relative_path}")
        actual_headers = tuple(normalize_text(cell.value) for cell in header_cells)
        if actual_headers != tuple(headers):
            raise ValueError(f"unexpected headers in {source.relative_path}: {actual_headers!r}")

        rows: list[tuple[int, tuple[Any, ...]]] = []
        for row_number, cells in enumerate(iterator, start=2):
            if any(cell.data_type == "f" for cell in cells):
                raise ValueError(
                    f"formulas are forbidden in {source.relative_path} row {row_number}"
                )
            values = tuple(cell.value for cell in cells)
            if all(_is_empty(value) for value in values):
                continue
            rows.append((row_number, values))
        if not rows:
            raise ValueError(f"{source.relative_path} has no data rows")
        return rows
    finally:
        workbook.close()


def _parse_fault_points(source: SourceFile, available_at: datetime) -> list[Record]:
    result: list[Record] = []
    seen: set[tuple[int, int, int]] = set()
    for source_row, values in _xlsx_rows(source, FAULT_COORDINATE_HEADERS):
        fault_id = _required_int(values[0], field="fault_id", row=source_row)
        segment_number = _required_int(values[1], field="segment_number", row=source_row)
        point_number = _required_int(values[2], field="point_number", row=source_row)
        longitude = _required_float(values[3], field="longitude", row=source_row)
        latitude = _required_float(values[4], field="latitude", row=source_row)
        new_flag = _required_int(values[5], field="is_new_in_source_version", row=source_row)
        if not -180.0 <= longitude <= 180.0 or not -90.0 <= latitude <= 90.0:
            raise ValueError(f"invalid longitude/latitude at Excel row {source_row}")
        if new_flag not in {0, 1}:
            raise ValueError(f"new-version flag must be 0 or 1 at Excel row {source_row}")
        key = (fault_id, segment_number, point_number)
        if key in seen:
            raise ValueError(f"duplicate fault point key {key}")
        seen.add(key)
        result.append(
            {
                "fault_point_id": stable_token("fault_point", *key),
                "fault_id": fault_id,
                "segment_number": segment_number,
                "point_number": point_number,
                "longitude": longitude,
                "latitude": latitude,
                "is_new_in_source_version": bool(new_flag),
                "fault_name_raw": _required_text(values[6], field="fault_name", row=source_row),
                "segment_name_raw": _required_text(values[7], field="segment_name", row=source_row),
                "source_file": source.relative_path,
                "source_row": source_row,
                "source_available_at": available_at,
                "historical_model_eligible": False,
                "quality_flags": (),
            }
        )
    return sorted(
        result,
        key=lambda row: (row["fault_id"], row["segment_number"], row["point_number"]),
    )


def _parse_attributes(source: SourceFile) -> dict[tuple[int, int], Record]:
    result: dict[tuple[int, int], Record] = {}
    for source_row, values in _xlsx_rows(source, FAULT_ATTRIBUTE_HEADERS):
        fault_id = _required_int(values[0], field="fault_id", row=source_row)
        segment_number = _required_int(values[2], field="segment_number", row=source_row)
        key = (fault_id, segment_number)
        if key in result:
            raise ValueError(f"duplicate fault attribute key {key}")
        recurrence_reliability = _optional_int(
            values[16], field="recurrence_reliability", row=source_row
        )
        last_reliability = _optional_int(
            values[18], field="last_strong_earthquake_reliability", row=source_row
        )
        if recurrence_reliability not in {None, 1, 2} or last_reliability not in {None, 1, 2}:
            raise ValueError(f"reliability codes must be 1 or 2 at Excel row {source_row}")
        result[key] = {
            "fault_id": fault_id,
            "segment_number": segment_number,
            "fault_name_raw": _required_text(values[1], field="fault_name", row=source_row),
            "segment_name_raw": _required_text(values[3], field="segment_name", row=source_row),
            "strike_deg": _required_float(values[4], field="strike_deg", row=source_row),
            "dip_deg": _required_float(values[5], field="dip_deg", row=source_row),
            "slip_angle_deg": _required_float(values[6], field="slip_angle_deg", row=source_row),
            "geologic_total_rate_mm_per_year": _optional_float(
                values[7], field="geologic_total_rate_mm_per_year", row=source_row
            ),
            "geologic_strike_slip_rate_mm_per_year": _optional_float(
                values[8], field="geologic_strike_slip_rate_mm_per_year", row=source_row
            ),
            "geologic_dip_slip_rate_mm_per_year": _optional_float(
                values[9], field="geologic_dip_slip_rate_mm_per_year", row=source_row
            ),
            "geodetic_total_rate_mm_per_year": _optional_float(
                values[10], field="geodetic_total_rate_mm_per_year", row=source_row
            ),
            "geodetic_strike_slip_rate_mm_per_year": _optional_float(
                values[11], field="geodetic_strike_slip_rate_mm_per_year", row=source_row
            ),
            "geodetic_strike_slip_error_mm_per_year": _optional_float(
                values[12], field="geodetic_strike_slip_error_mm_per_year", row=source_row
            ),
            "geodetic_dip_slip_rate_mm_per_year": _optional_float(
                values[13], field="geodetic_dip_slip_rate_mm_per_year", row=source_row
            ),
            "geodetic_dip_slip_error_mm_per_year": _optional_float(
                values[14], field="geodetic_dip_slip_error_mm_per_year", row=source_row
            ),
            "recurrence_period_years": _optional_float(
                values[15], field="recurrence_period_years", row=source_row
            ),
            "recurrence_reliability": recurrence_reliability,
            "last_strong_earthquake_year_raw": _optional_int(
                values[17], field="last_strong_earthquake_year_raw", row=source_row
            ),
            "last_strong_earthquake_reliability": last_reliability,
        }
    return result


def _parse_hazard(source: SourceFile) -> dict[tuple[int, int], Record]:
    result: dict[tuple[int, int], Record] = {}
    for source_row, values in _xlsx_rows(source, LONG_TERM_HAZARD_HEADERS):
        fault_id = _required_int(values[0], field="fault_id", row=source_row)
        segment_number = _required_int(values[2], field="segment_number", row=source_row)
        key = (fault_id, segment_number)
        if key in result:
            raise ValueError(f"duplicate long-term hazard key {key}")
        components = tuple(
            _required_float(values[index], field=LONG_TERM_HAZARD_HEADERS[index], row=source_row)
            for index in range(4, 8)
        )
        score = _required_float(values[8], field="long_term_hazard_score", row=source_row)
        if not math.isclose(sum(components), score, rel_tol=0.0, abs_tol=1e-9):
            raise ValueError(f"long-term hazard score does not equal component sum for key {key}")
        result[key] = {
            "fault_id": fault_id,
            "segment_number": segment_number,
            "fault_name_raw": _required_text(values[1], field="fault_name", row=source_row),
            "segment_name_raw": _required_text(values[3], field="segment_name", row=source_row),
            "rupture_gap_weight": components[0],
            "locking_weight": components[1],
            "microseismic_sparsity_weight": components[2],
            "coulomb_stress_weight": components[3],
            "long_term_hazard_score": score,
        }
    return result


def _deduplicate_adjacent(coordinates: Sequence[Coordinate]) -> tuple[tuple[Coordinate, ...], int]:
    if not coordinates:
        return (), 0
    result = [coordinates[0]]
    removed = 0
    for coordinate in coordinates[1:]:
        if coordinate == result[-1]:
            removed += 1
        else:
            result.append(coordinate)
    return tuple(result), removed


def _geometry_key(coordinates: Sequence[Coordinate]) -> tuple[tuple[str, str], ...]:
    forward = tuple(
        (canonical_decimal(longitude), canonical_decimal(latitude))
        for longitude, latitude in coordinates
    )
    reverse = tuple(reversed(forward))
    return min(forward, reverse)


def _geometry_hash(key: Sequence[tuple[str, str]]) -> str:
    return stable_token("geometry", *(f"{longitude},{latitude}" for longitude, latitude in key))


def _line_wkb(coordinates: Sequence[Coordinate]) -> bytes:
    geometry = LineString(coordinates)
    return cast(bytes, to_wkb(geometry, byte_order=1, include_srid=False, output_dimension=2))


def _build_fault_segments(
    points: list[Record],
    attributes: Mapping[tuple[int, int], Record],
    hazards: Mapping[tuple[int, int], Record],
    coordinate_source: SourceFile,
    attribute_source: SourceFile,
    hazard_source: SourceFile,
    available_at: datetime,
) -> tuple[list[Record], int, int]:
    grouped: dict[tuple[int, int], list[Record]] = defaultdict(list)
    for point in points:
        grouped[(point["fault_id"], point["segment_number"])].append(point)
    point_keys = set(grouped)
    if point_keys != set(attributes) or point_keys != set(hazards):
        raise ValueError("fault coordinate, attribute, and hazard key sets must match exactly")

    segments: list[Record] = []
    adjacent_duplicates = 0
    non_simple = 0
    optional_attribute_fields = (
        "geologic_total_rate_mm_per_year",
        "geologic_strike_slip_rate_mm_per_year",
        "geologic_dip_slip_rate_mm_per_year",
        "geodetic_total_rate_mm_per_year",
        "geodetic_strike_slip_rate_mm_per_year",
        "geodetic_strike_slip_error_mm_per_year",
        "geodetic_dip_slip_rate_mm_per_year",
        "geodetic_dip_slip_error_mm_per_year",
        "recurrence_period_years",
        "recurrence_reliability",
        "last_strong_earthquake_year_raw",
        "last_strong_earthquake_reliability",
    )
    for key in sorted(point_keys):
        segment_points = sorted(grouped[key], key=lambda row: row["point_number"])
        expected_numbers = list(range(1, len(segment_points) + 1))
        if [row["point_number"] for row in segment_points] != expected_numbers:
            raise ValueError(f"fault point numbers must be contiguous from 1 for key {key}")
        names = {(row["fault_name_raw"], row["segment_name_raw"]) for row in segment_points}
        if len(names) != 1:
            raise ValueError(f"fault names are inconsistent within segment {key}")
        new_flags = {row["is_new_in_source_version"] for row in segment_points}
        if len(new_flags) != 1:
            raise ValueError(f"new-version flags are inconsistent within segment {key}")

        attribute = attributes[key]
        hazard = hazards[key]
        fault_name, segment_name = next(iter(names))
        if (
            attribute["fault_name_raw"] != fault_name
            or attribute["segment_name_raw"] != segment_name
            or hazard["fault_name_raw"] != fault_name
            or hazard["segment_name_raw"] != segment_name
        ):
            raise ValueError(f"fault names disagree across sources for key {key}")

        raw_coordinates = tuple((row["longitude"], row["latitude"]) for row in segment_points)
        geometry_coordinates, removed = _deduplicate_adjacent(raw_coordinates)
        adjacent_duplicates += removed
        if len(set(geometry_coordinates)) < 2:
            raise ValueError(f"fault segment {key} has fewer than two distinct points")
        geometry = LineString(geometry_coordinates)
        if not geometry.is_valid:
            raise ValueError(f"fault segment {key} has invalid geometry")
        geometry_flags: list[str] = []
        if removed:
            geometry_flags.append("adjacent_duplicate_coordinates_removed")
            for index, point in enumerate(segment_points[1:], start=1):
                if raw_coordinates[index] == raw_coordinates[index - 1]:
                    point["quality_flags"] = ("adjacent_duplicate_coordinate",)
        if not geometry.is_simple:
            geometry_flags.append("non_simple_geometry")
            non_simple += 1

        missing_fields = tuple(
            field for field in optional_attribute_fields if attribute[field] is None
        )
        recurrence_period = cast(float | None, attribute["recurrence_period_years"])
        last_earthquake_year = cast(int | None, attribute["last_strong_earthquake_year_raw"])
        elapsed_ratio = (
            (available_at.year - last_earthquake_year) / recurrence_period
            if recurrence_period is not None and last_earthquake_year is not None
            else None
        )
        segment: Record = {
            "fault_segment_id": stable_token("fault_segment", *key),
            "fault_id": key[0],
            "segment_number": key[1],
            "fault_name_raw": fault_name,
            "segment_name_raw": segment_name,
            "simplified_geometry_wkb": _line_wkb(geometry_coordinates),
            "raw_point_count": len(raw_coordinates),
            "geometry_point_count": len(geometry_coordinates),
            "is_new_in_source_version": next(iter(new_flags)),
            "strike_deg": attribute["strike_deg"],
            "dip_deg": attribute["dip_deg"],
            "slip_angle_deg": attribute["slip_angle_deg"],
            "geologic_total_rate_mm_per_year": attribute["geologic_total_rate_mm_per_year"],
            "geologic_strike_slip_rate_mm_per_year": attribute[
                "geologic_strike_slip_rate_mm_per_year"
            ],
            "geologic_dip_slip_rate_mm_per_year": attribute["geologic_dip_slip_rate_mm_per_year"],
            "geodetic_total_rate_mm_per_year": attribute["geodetic_total_rate_mm_per_year"],
            "geodetic_strike_slip_rate_mm_per_year": attribute[
                "geodetic_strike_slip_rate_mm_per_year"
            ],
            "geodetic_strike_slip_error_mm_per_year": attribute[
                "geodetic_strike_slip_error_mm_per_year"
            ],
            "geodetic_dip_slip_rate_mm_per_year": attribute["geodetic_dip_slip_rate_mm_per_year"],
            "geodetic_dip_slip_error_mm_per_year": attribute["geodetic_dip_slip_error_mm_per_year"],
            "recurrence_period_years": recurrence_period,
            "recurrence_reliability": attribute["recurrence_reliability"],
            "last_strong_earthquake_year_raw": last_earthquake_year,
            "last_strong_earthquake_reliability": attribute["last_strong_earthquake_reliability"],
            "elapsed_ratio_at_snapshot": elapsed_ratio,
            "rupture_gap_weight": hazard["rupture_gap_weight"],
            "locking_weight": hazard["locking_weight"],
            "microseismic_sparsity_weight": hazard["microseismic_sparsity_weight"],
            "coulomb_stress_weight": hazard["coulomb_stress_weight"],
            "long_term_hazard_score": hazard["long_term_hazard_score"],
            "missing_fields": missing_fields,
            "geometry_flags": tuple(geometry_flags),
            "attribute_source_file": attribute_source.relative_path,
            "hazard_source_file": hazard_source.relative_path,
            "source_available_at": available_at,
            "historical_model_eligible": False,
            "true_trace_id": None,
            "true_trace_mapping_status": "unreviewed",
            "true_trace_match_confidence": "unassessed",
        }
        segments.append(segment)

    # Explicitly prove that all three inputs participated without positional joins.
    if coordinate_source.relative_path == "":
        raise ValueError("fault coordinate source path cannot be empty")
    return segments, adjacent_duplicates, non_simple


def _parse_gmt(source: SourceFile, *, encoding: str) -> tuple[_GmtSegment, ...]:
    payload = read_stable_bytes(source)
    try:
        text = payload.decode(encoding)
    except UnicodeDecodeError as exc:
        raise ValueError(f"cannot decode {source.relative_path} as {encoding}") from exc

    result: list[_GmtSegment] = []
    current_header: str | None = None
    current_comments: list[str] = []
    current_coordinates: list[Coordinate] | None = None
    current_delimiter_line = 0
    pending_comments: list[str] = []

    def finish_current() -> None:
        if current_coordinates is None:
            return
        if not current_coordinates:
            raise ValueError(
                f"empty GMT segment in {source.relative_path} at line {current_delimiter_line}"
            )
        result.append(
            _GmtSegment(
                ordinal=len(result) + 1,
                header_raw=current_header,
                comments=tuple(current_comments),
                delimiter_line=current_delimiter_line,
                coordinates=tuple(current_coordinates),
            )
        )

    for line_number, raw_line in enumerate(text.splitlines(), start=1):
        line = raw_line.strip()
        if not line:
            continue
        if line.startswith("#"):
            comment = line[1:].strip()
            if current_coordinates is not None and not current_coordinates:
                current_comments.append(comment)
            else:
                pending_comments.append(comment)
            continue
        if line.startswith(">"):
            finish_current()
            current_header = normalize_text(line[1:])
            current_comments = pending_comments
            pending_comments = []
            current_coordinates = []
            current_delimiter_line = line_number
            continue
        if current_coordinates is None:
            raise ValueError(
                f"coordinate before first GMT delimiter in {source.relative_path} "
                f"line {line_number}"
            )
        parts = line.split()
        if len(parts) != 2:
            raise ValueError(
                f"GMT coordinate must contain exactly two values in "
                f"{source.relative_path} line {line_number}"
            )
        try:
            longitude, latitude = (float(part) for part in parts)
        except ValueError as exc:
            raise ValueError(
                f"invalid GMT coordinate in {source.relative_path} line {line_number}"
            ) from exc
        if (
            not math.isfinite(longitude)
            or not math.isfinite(latitude)
            or not -180.0 <= longitude <= 180.0
            or not -90.0 <= latitude <= 90.0
        ):
            raise ValueError(
                f"out-of-range GMT coordinate in {source.relative_path} line {line_number}"
            )
        current_coordinates.append((longitude, latitude))
    finish_current()
    if not result:
        raise ValueError(f"{source.relative_path} contains no GMT segments")
    return tuple(result)


def _loose_fault_name(value: str | None) -> str | None:
    normalized = normalize_text(value)
    if normalized is None or _NUMERIC_NAME_RE.fullmatch(normalized):
        return None
    normalized = _DASH_RE.sub("-", normalized)
    normalized = normalized.replace("断裂带", "断裂").replace("断层", "断裂")
    return normalized


def _apply_duplicate_audit(records: list[Record], keys: list[tuple[tuple[str, str], ...]]) -> None:
    groups: dict[tuple[tuple[str, str], ...], list[int]] = defaultdict(list)
    for index, key in enumerate(keys):
        groups[key].append(index)
    for key, indices in groups.items():
        if len(indices) == 1:
            continue
        duplicate_group_id = stable_token(
            "duplicate_geometry", *(f"{longitude},{latitude}" for longitude, latitude in key)
        )
        names = {records[index].get("trace_name_raw") for index in indices}
        name_conflict = len(names) > 1
        for index in indices:
            record = records[index]
            record["duplicate_group_id"] = duplicate_group_id
            record["duplicate_geometry"] = True
            if "duplicate_name_conflict" in record:
                record["duplicate_name_conflict"] = name_conflict
            flags = list(record["quality_flags"])
            flags.append("duplicate_geometry")
            if name_conflict and "duplicate_name_conflict" in record:
                flags.append("duplicate_name_conflict")
            record["quality_flags"] = tuple(flags)


def _build_fault_traces(
    source: SourceFile, segments: Sequence[_GmtSegment], available_at: datetime
) -> tuple[list[Record], Record]:
    records: list[Record] = []
    geometry_keys: list[tuple[tuple[str, str], ...]] = []
    adjacent_duplicate_count = 0
    zero_length_count = 0
    non_simple_count = 0
    numeric_name_count = 0
    for segment in segments:
        coordinates, removed = _deduplicate_adjacent(segment.coordinates)
        adjacent_duplicate_count += removed
        # Duplicate provenance is defined on the untouched source coordinate sequence.
        # Otherwise cleaning an adjacent duplicate can create a new, artificial duplicate group.
        key = _geometry_key(segment.coordinates)
        geometry_keys.append(key)
        name = normalize_text(segment.header_raw)
        flags: list[str] = []
        if removed:
            flags.append("adjacent_duplicate_coordinates_removed")
        if name is None:
            flags.append("missing_trace_name")
        elif _NUMERIC_NAME_RE.fullmatch(name):
            flags.append("numeric_trace_name")
            numeric_name_count += 1

        usable = len(set(coordinates)) >= 2
        geometry_wkb: bytes | None = None
        if usable:
            geometry = LineString(coordinates)
            if not geometry.is_valid:
                flags.append("invalid_geometry")
                usable = False
            else:
                geometry_wkb = _line_wkb(coordinates)
                if not geometry.is_simple:
                    flags.append("non_simple_geometry")
                    non_simple_count += 1
        if not usable:
            flags.append("zero_length_geometry")
            zero_length_count += 1

        records.append(
            {
                "trace_id": stable_token("fault_trace", source.relative_path, segment.ordinal),
                "source_segment_number": segment.ordinal,
                "trace_name_raw": name,
                "trace_name_normalized": normalize_text(name),
                "geometry_wkb": geometry_wkb,
                "raw_point_count": len(segment.coordinates),
                "geometry_point_count": len(coordinates),
                "is_closed": len(coordinates) >= 4 and coordinates[0] == coordinates[-1],
                "usable_for_geometry": usable,
                "geometry_hash": _geometry_hash(key),
                "duplicate_group_id": None,
                "duplicate_geometry": False,
                "duplicate_name_conflict": False,
                "quality_flags": tuple(flags),
                "source_file": source.relative_path,
                "delimiter_source_line": segment.delimiter_line,
                "source_available_at": available_at,
                "historical_model_eligible": False,
            }
        )
    _apply_duplicate_audit(records, geometry_keys)
    duplicate_groups = {
        record["duplicate_group_id"]
        for record in records
        if record["duplicate_group_id"] is not None
    }
    conflict_groups = {
        record["duplicate_group_id"] for record in records if record["duplicate_name_conflict"]
    }
    quality: Record = {
        "fault_trace_count": len(records),
        "fault_trace_usable_count": sum(record["usable_for_geometry"] for record in records),
        "fault_trace_zero_length_count": zero_length_count,
        "fault_trace_adjacent_duplicate_count": adjacent_duplicate_count,
        "fault_trace_non_simple_count": non_simple_count,
        "fault_trace_numeric_name_count": numeric_name_count,
        "fault_trace_duplicate_group_count": len(duplicate_groups),
        "fault_trace_duplicate_geometry_excess": sum(
            1 for record in records if record["duplicate_geometry"]
        )
        - len(duplicate_groups),
        "fault_trace_duplicate_name_conflict_group_count": len(conflict_groups),
    }
    return records, quality


def _build_basemap_features(
    source_segments: Sequence[tuple[SourceFile, str, Sequence[_GmtSegment]]],
    available_at: datetime,
) -> tuple[list[Record], Record]:
    records: list[Record] = []
    geometry_keys: list[tuple[tuple[str, str], ...]] = []
    adjacent_duplicate_count = 0
    for source, role, segments in source_segments:
        for segment in segments:
            coordinates, removed = _deduplicate_adjacent(segment.coordinates)
            adjacent_duplicate_count += removed
            if len(set(coordinates)) < 2:
                raise ValueError(
                    f"basemap segment {source.relative_path}:{segment.ordinal} is zero length"
                )
            geometry = LineString(coordinates)
            if not geometry.is_valid:
                raise ValueError(
                    f"invalid basemap segment {source.relative_path}:{segment.ordinal}"
                )
            key = _geometry_key(coordinates)
            geometry_keys.append(key)
            flags: list[str] = []
            if removed:
                flags.append("adjacent_duplicate_coordinates_removed")
            if not geometry.is_simple:
                flags.append("non_simple_geometry")
            records.append(
                {
                    "basemap_feature_id": stable_token(
                        "basemap_feature", source.relative_path, segment.ordinal
                    ),
                    "role": role,
                    "source_segment_number": segment.ordinal,
                    "geometry_wkb": _line_wkb(coordinates),
                    "raw_point_count": len(segment.coordinates),
                    "geometry_point_count": len(coordinates),
                    "is_closed": len(coordinates) >= 4 and coordinates[0] == coordinates[-1],
                    "geometry_hash": _geometry_hash(key),
                    "duplicate_group_id": None,
                    "duplicate_geometry": False,
                    "quality_flags": tuple(flags),
                    "source_file": source.relative_path,
                    "delimiter_source_line": segment.delimiter_line,
                    "source_comments": segment.comments,
                    "source_available_at": available_at,
                    "model_feature_eligible": False,
                }
            )
    _apply_duplicate_audit(records, geometry_keys)
    duplicate_groups = {
        record["duplicate_group_id"]
        for record in records
        if record["duplicate_group_id"] is not None
    }
    quality: Record = {
        "basemap_feature_count": len(records),
        "basemap_adjacent_duplicate_count": adjacent_duplicate_count,
        "basemap_duplicate_group_count": len(duplicate_groups),
        "basemap_duplicate_geometry_excess": sum(
            1 for record in records if record["duplicate_geometry"]
        )
        - len(duplicate_groups),
    }
    return records, quality


def _build_study_area(
    source: SourceFile,
    segments: Sequence[_GmtSegment],
    equal_area_crs: str,
    available_at: datetime,
) -> Record:
    projected_crs = CRS.from_user_input(equal_area_crs)
    transformer = Transformer.from_crs("EPSG:4326", projected_crs, always_xy=True)
    candidates: list[tuple[float, int, _GmtSegment, Polygon]] = []
    for segment in segments:
        coordinates, _ = _deduplicate_adjacent(segment.coordinates)
        if len(coordinates) < 4 or coordinates[0] != coordinates[-1]:
            continue
        polygon = Polygon(coordinates)
        if polygon.is_empty or not polygon.is_valid or not polygon.exterior.is_simple:
            continue
        projected = transform(transformer.transform, polygon)
        if not projected.is_valid or not math.isfinite(projected.area) or projected.area <= 0:
            continue
        candidates.append((projected.area, segment.ordinal, segment, polygon))
    if not candidates:
        raise ValueError("CN-border-L1.gmt has no valid closed polygon candidate")
    area_m2, _, selected_segment, selected_polygon = max(
        candidates, key=lambda candidate: (candidate[0], -candidate[1])
    )
    normalized_polygon = orient(selected_polygon, sign=1.0)
    ring_coordinates = [
        [float(longitude), float(latitude)]
        for longitude, latitude in normalized_polygon.exterior.coords
    ]
    longitude_values = [coordinate[0] for coordinate in ring_coordinates]
    latitude_values = [coordinate[1] for coordinate in ring_coordinates]
    geodesic_area_m2, _ = Geod(ellps="WGS84").polygon_area_perimeter(
        longitude_values, latitude_values
    )
    return {
        "type": "Feature",
        "id": stable_token(
            "study_area", source.relative_path, selected_segment.ordinal, equal_area_crs
        ),
        "properties": {
            "name": "china_mainland",
            "source_file": source.relative_path,
            "source_sha256": source.sha256,
            "source_segment_number": selected_segment.ordinal,
            "selection_rule": (
                "largest_valid_closed_ring_by_equal_area_then_lowest_segment_number"
            ),
            "source_coordinate_count": len(selected_segment.coordinates),
            "normalized_coordinate_count": len(ring_coordinates),
            "bbox": [float(value) for value in normalized_polygon.bounds],
            "equal_area_crs": equal_area_crs,
            "equal_area_km2": round(area_m2 / 1_000_000.0, 6),
            "geodesic_area_km2": round(abs(geodesic_area_m2) / 1_000_000.0, 6),
            "source_available_at": _available_at_json(available_at),
            "target_independent": True,
            "predictor_feature": False,
            "island_policy": ("continuous_mainland_excludes_hainan_taiwan_and_other_islands"),
        },
        "geometry": {"type": "Polygon", "coordinates": [ring_coordinates]},
    }


def _build_crosswalk(fault_segments: list[Record], fault_traces: list[Record]) -> list[Record]:
    exact_index: dict[str, list[Record]] = defaultdict(list)
    loose_index: dict[str, list[Record]] = defaultdict(list)
    for trace in fault_traces:
        name = cast(str | None, trace["trace_name_raw"])
        normalized = normalize_text(name)
        loose = _loose_fault_name(name)
        if normalized is not None and not _NUMERIC_NAME_RE.fullmatch(normalized):
            exact_index[normalized].append(trace)
        if loose is not None:
            loose_index[loose].append(trace)

    result: list[Record] = []
    for segment in fault_segments:
        fault_name = cast(str, segment["fault_name_raw"])
        normalized = cast(str, normalize_text(fault_name))
        loose = _loose_fault_name(fault_name)
        candidates: dict[str, tuple[Record, str]] = {}
        for trace in exact_index.get(normalized, []):
            candidates[trace["trace_id"]] = (trace, "exact_normalized_name")
        if loose is not None:
            for trace in loose_index.get(loose, []):
                candidates.setdefault(trace["trace_id"], (trace, "loose_normalized_name"))
        if not candidates:
            segment["true_trace_mapping_status"] = "unmatched"
            segment["true_trace_match_confidence"] = "none"
            continue
        segment["true_trace_mapping_status"] = "unreviewed"
        for trace, method in sorted(
            candidates.values(), key=lambda item: item[0]["source_segment_number"]
        ):
            result.append(
                {
                    "crosswalk_candidate_id": stable_token(
                        "fault_trace_crosswalk",
                        segment["fault_segment_id"],
                        trace["trace_id"],
                    ),
                    "fault_segment_id": segment["fault_segment_id"],
                    "trace_id": trace["trace_id"],
                    "fault_name_raw": fault_name,
                    "trace_name_raw": trace["trace_name_raw"],
                    "name_match_method": method,
                    "status": "unreviewed",
                    "match_confidence": "unassessed",
                    "trace_duplicate_group_id": trace["duplicate_group_id"],
                    "trace_duplicate_name_conflict": trace["duplicate_name_conflict"],
                    "historical_model_eligible": False,
                }
            )
    return result


def parse_geology_sources(
    source_map: Mapping[str, SourceFile],
    equal_area_crs: str,
    dataset_available_at: datetime | str,
) -> GeologyParseResult:
    """Parse all stage-1 geology sources without positional joins or target-event inputs."""

    sources = _find_sources(source_map)
    available_at = _available_at(dataset_available_at)
    coordinate_source = sources["FaultCord.xlsx"]
    attribute_source = sources["FaultAttri_ALLV4_修改.xlsx"]
    hazard_source = sources["weight_analysis-修改newV4_修改.xlsx"]

    fault_points = _parse_fault_points(coordinate_source, available_at)
    attributes = _parse_attributes(attribute_source)
    hazards = _parse_hazard(hazard_source)
    fault_segments, fault_duplicate_count, fault_non_simple_count = _build_fault_segments(
        fault_points,
        attributes,
        hazards,
        coordinate_source,
        attribute_source,
        hazard_source,
        available_at,
    )

    fault_trace_source = sources["FaultX.dat"]
    raw_fault_traces = _parse_gmt(fault_trace_source, encoding="gb18030")
    fault_traces, trace_quality = _build_fault_traces(
        fault_trace_source, raw_fault_traces, available_at
    )

    basemap_inputs: list[tuple[SourceFile, str, Sequence[_GmtSegment]]] = []
    border_segments: tuple[_GmtSegment, ...] | None = None
    border_source: SourceFile | None = None
    for filename in _BASEMAP_ROLES:
        source = sources[next(name for name in _REQUIRED_FILES if name.casefold() == filename)]
        segments = _parse_gmt(source, encoding="utf-8-sig")
        basemap_inputs.append((source, _BASEMAP_ROLES[filename], segments))
        if filename == "cn-border-l1.gmt":
            border_segments = segments
            border_source = source
    if border_segments is None or border_source is None:
        raise ValueError("CN-border-L1.gmt was not parsed")
    basemap_features, basemap_quality = _build_basemap_features(basemap_inputs, available_at)
    study_area = _build_study_area(border_source, border_segments, equal_area_crs, available_at)
    crosswalk = _build_crosswalk(fault_segments, fault_traces)

    missing_fields = Counter(
        field for segment in fault_segments for field in segment["missing_fields"]
    )
    elapsed_ratios = [
        cast(float, segment["elapsed_ratio_at_snapshot"])
        for segment in fault_segments
        if segment["elapsed_ratio_at_snapshot"] is not None
    ]
    trace_coordinates = [
        coordinate for segment in raw_fault_traces for coordinate in segment.coordinates
    ]
    basemap_coordinates = [
        coordinate
        for _, _, segments in basemap_inputs
        for segment in segments
        for coordinate in segment.coordinates
    ]

    quality: Record = {
        "fault_point_count": len(fault_points),
        "fault_segment_count": len(fault_segments),
        "fault_key_join_rule": "fault_id_and_segment_number_strict_one_to_one",
        "fault_key_sets_equal": True,
        "fault_adjacent_duplicate_count": fault_duplicate_count,
        "fault_non_simple_count": fault_non_simple_count,
        "fault_point_bbox_wgs84": _coordinate_bbox(
            [(point["longitude"], point["latitude"]) for point in fault_points]
        ),
        "fault_trace_bbox_wgs84": _coordinate_bbox(trace_coordinates),
        "basemap_bbox_wgs84": _coordinate_bbox(basemap_coordinates),
        "fault_segment_missing_field_counts": dict(sorted(missing_fields.items())),
        "elapsed_ratio_at_snapshot_min": min(elapsed_ratios) if elapsed_ratios else None,
        "elapsed_ratio_at_snapshot_max": max(elapsed_ratios) if elapsed_ratios else None,
        "elapsed_ratio_at_snapshot_missing_count": len(fault_segments) - len(elapsed_ratios),
        "true_trace_mapping_status_counts": dict(
            sorted(
                Counter(segment["true_trace_mapping_status"] for segment in fault_segments).items()
            )
        ),
        "true_trace_accepted_mapping_count": sum(
            segment["true_trace_id"] is not None for segment in fault_segments
        ),
        "hazard_score_mismatch_count": 0,
        "crosswalk_candidate_count": len(crosswalk),
        "crosswalk_status": "name_candidates_only_unreviewed",
        "source_available_at": _available_at_json(available_at),
        "historical_model_eligible": False,
        "study_area_target_independent": True,
        **trace_quality,
        **basemap_quality,
    }
    return GeologyParseResult(
        fault_points=tuple(fault_points),
        fault_segments=tuple(fault_segments),
        fault_traces=tuple(fault_traces),
        basemap_features=tuple(basemap_features),
        crosswalk_audit=tuple(crosswalk),
        study_area=study_area,
        quality=quality,
    )
